from io import BytesIO

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import DetailView, FormView, ListView, View

from apps.projects.models import Project
from apps.projects.utils import is_valid_domain, normalize_domain
from .forms import ImportExcelForm
from .models import ImportRow, ImportSession


class CanImportMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_admin_role or self.request.user.is_manager_role


class ImportListView(CanImportMixin, ListView):
    model = ImportSession
    template_name = "imports/import_list.html"

    def get_queryset(self):
        qs = super().get_queryset()
        if self.request.user.is_manager_role:
            return qs.filter(imported_by=self.request.user)
        return qs


class ImportCreateView(CanImportMixin, FormView):
    form_class = ImportExcelForm
    template_name = "imports/import_form.html"
    success_url = reverse_lazy("imports")

    def form_valid(self, form):
        from openpyxl import load_workbook

        file = form.cleaned_data["file"]
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        session = ImportSession.objects.create(imported_by=self.request.user, original_file_name=file.name)
        seen = set()
        total = valid = errors = dup_file = dup_system = 0
        existing_map = {p.domain: p for p in Project.objects.all()}
        for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            raw = str(row[0] or "").strip()
            if not raw or raw.lower() in {"link", "domain"}:
                continue
            total += 1
            domain = normalize_domain(raw)
            row_data = {"session": session, "row_number": index, "raw_value": raw, "domain": domain}
            if not is_valid_domain(domain):
                errors += 1
                row_data["error_message"] = "Domain không hợp lệ"
            elif domain in seen:
                dup_file += 1
                row_data.update({"is_valid": True, "is_duplicate_file": True, "error_message": "Trùng trong file"})
            elif domain in existing_map:
                dup_system += 1
                row_data.update({"is_valid": True, "is_duplicate_system": True, "existing_project_id": existing_map[domain].id, "error_message": "Đã tồn tại trong hệ thống"})
            else:
                valid += 1
                row_data["is_valid"] = True
            seen.add(domain)
            ImportRow.objects.create(**row_data)
        session.total_rows = total
        session.valid_rows = valid
        session.error_rows = errors
        session.duplicate_in_file = dup_file
        session.duplicate_in_system = dup_system
        session.save()
        return redirect("import_detail", pk=session.pk)


class ImportDetailView(CanImportMixin, DetailView):
    model = ImportSession
    template_name = "imports/import_detail.html"

    def get_queryset(self):
        qs = super().get_queryset()
        if self.request.user.is_manager_role:
            return qs.filter(imported_by=self.request.user)
        return qs


class ImportCommitView(CanImportMixin, View):
    def post(self, request, pk):
        sessions = ImportSession.objects.filter(committed=False)
        if request.user.is_manager_role:
            sessions = sessions.filter(imported_by=request.user)
        session = get_object_or_404(sessions, pk=pk)
        count = 0
        rows = session.rows.filter(is_valid=True, is_duplicate_file=False, is_duplicate_system=False)
        for row in rows:
            data = {"domain": row.domain, "created_by": request.user}
            if request.user.is_manager_role:
                data.update({"manager": request.user, "status": Project.Status.UNASSIGNED})
            project = Project.objects.create(**data)
            row.imported_project_id = project.id
            row.save(update_fields=["imported_project_id"])
            count += 1
        session.imported_success = count
        session.committed = True
        session.save(update_fields=["imported_success", "committed"])
        messages.success(request, f"Đã import {count} domain hợp lệ.")
        return redirect("import_detail", pk=session.pk)


class ImportReportView(CanImportMixin, View):
    def get(self, request, pk):
        sessions = ImportSession.objects.all()
        if request.user.is_manager_role:
            sessions = sessions.filter(imported_by=request.user)
        session = get_object_or_404(sessions, pk=pk)
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Bao cao import"
        ws.append(["Dòng", "Giá trị gốc", "Domain", "Hợp lệ", "Trùng file", "Trùng hệ thống", "Lỗi"])
        for row in session.rows.all():
            ws.append([row.row_number, row.raw_value, row.domain, row.is_valid, row.is_duplicate_file, row.is_duplicate_system, row.error_message])
        output = BytesIO()
        wb.save(output)
        response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="import-{session.pk}.xlsx"'
        return response
