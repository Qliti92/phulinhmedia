from django.test import TestCase
from django.urls import reverse
from io import BytesIO

from openpyxl import load_workbook

from apps.accounts.models import User
from .models import Project


class ProjectDeletePermissionTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            email="admin@example.com", password="test", full_name="Admin", role=User.Role.ADMIN
        )
        self.manager = User.objects.create_user(
            email="manager@example.com", password="test", full_name="Manager", role=User.Role.MANAGER
        )
        self.other_manager = User.objects.create_user(
            email="other@example.com", password="test", full_name="Other", role=User.Role.MANAGER
        )
        self.staff = User.objects.create_user(
            email="staff@example.com", password="test", full_name="Staff", role=User.Role.STAFF
        )

    def test_admin_can_delete_any_project(self):
        project = Project.objects.create(domain="admin-delete.example", created_by=self.other_manager)
        self.client.force_login(self.admin)
        response = self.client.post(reverse("project_delete", args=[project.pk]))
        self.assertRedirects(response, reverse("projects"))
        self.assertFalse(Project.objects.filter(pk=project.pk).exists())

    def test_manager_can_delete_project_they_created(self):
        project = Project.objects.create(
            domain="manager-delete.example", manager=self.manager, created_by=self.manager
        )
        self.client.force_login(self.manager)
        response = self.client.post(reverse("project_delete", args=[project.pk]))
        self.assertRedirects(response, reverse("projects"))
        self.assertFalse(Project.objects.filter(pk=project.pk).exists())

    def test_manager_cannot_delete_project_created_by_someone_else(self):
        project = Project.objects.create(
            domain="other-manager.example", manager=self.manager, created_by=self.other_manager
        )
        self.client.force_login(self.manager)
        response = self.client.post(reverse("project_delete", args=[project.pk]))
        self.assertEqual(response.status_code, 403)
        self.assertTrue(Project.objects.filter(pk=project.pk).exists())

    def test_staff_cannot_delete_project(self):
        project = Project.objects.create(domain="staff-forbidden.example", staff=self.staff, created_by=self.staff)
        self.client.force_login(self.staff)
        response = self.client.post(reverse("project_delete", args=[project.pk]))
        self.assertEqual(response.status_code, 403)
        self.assertTrue(Project.objects.filter(pk=project.pk).exists())

    def test_admin_can_bulk_delete_projects(self):
        first = Project.objects.create(domain="bulk-one.example", created_by=self.manager)
        second = Project.objects.create(domain="bulk-two.example", created_by=self.other_manager)
        self.client.force_login(self.admin)
        response = self.client.post(reverse("project_bulk_delete"), {"project_ids": f"{first.pk},{second.pk}"})
        self.assertRedirects(response, reverse("projects"))
        self.assertFalse(Project.objects.filter(pk__in=[first.pk, second.pk]).exists())

    def test_staff_can_use_full_update_for_assigned_project(self):
        project = Project.objects.create(domain="staff-edit.example", staff=self.staff, created_by=self.manager)
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse("project_update", args=[project.pk]),
            {"status": Project.Status.IN_PROGRESS, "progress_stage": "", "financial_result": "profit"},
        )
        self.assertRedirects(response, reverse("projects"))
        project.refresh_from_db()
        self.assertEqual(project.financial_result, Project.FinancialResult.PROFIT)

    def test_project_export_supports_timezone_and_financial_columns(self):
        Project.objects.create(
            domain="export.example", created_by=self.admin, financial_result=Project.FinancialResult.LOSS
        )
        self.client.force_login(self.admin)
        response = self.client.get(reverse("export_projects"))
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.active.cell(row=2, column=8).value, "Lỗ")
