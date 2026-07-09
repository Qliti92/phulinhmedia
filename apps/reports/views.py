from io import BytesIO

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from django.views import View
from django.views.generic import TemplateView

from apps.core.models import ActivityLog
from apps.core.models import SystemSetting
from apps.projects.models import Project
from apps.tasks.models import Attendance, StaffPerformance, Task


def refresh_staff_performance():
    from apps.accounts.models import User

    for staff in User.objects.filter(role=User.Role.STAFF, is_active=True):
        item, _ = StaffPerformance.objects.get_or_create(staff=staff)
        item.recalculate()


def report_period(request):
    today = timezone.localdate()
    period = request.GET.get("period", "week")
    year = int(request.GET.get("year") or today.year)
    quarter = int(request.GET.get("quarter") or ((today.month - 1) // 3 + 1))
    if period == "quarter":
        start_month = (quarter - 1) * 3 + 1
        start = today.replace(year=year, month=start_month, day=1)
        end_month = start_month + 2
        next_month = today.replace(year=year + (1 if end_month == 12 else 0), month=1 if end_month == 12 else end_month + 1, day=1)
        end = next_month - timezone.timedelta(days=1)
        label = f"Quý {quarter}/{year}"
    elif period == "year":
        start = today.replace(year=year, month=1, day=1)
        end = today.replace(year=year, month=12, day=31)
        label = f"Năm {year}"
    elif period == "month":
        month = int(request.GET.get("month") or today.month)
        start = today.replace(year=year, month=month, day=1)
        next_month = today.replace(year=year + (1 if month == 12 else 0), month=1 if month == 12 else month + 1, day=1)
        end = next_month - timezone.timedelta(days=1)
        label = f"Tháng {month}/{year}"
    else:
        start = today - timezone.timedelta(days=today.weekday())
        end = start + timezone.timedelta(days=6)
        label = f"Tuần {start.strftime('%d/%m')} - {end.strftime('%d/%m/%Y')}"
    return period, start, end, label


class ReportDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "reports/report_dashboard.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        period, start, end, label = report_period(self.request)
        user = self.request.user
        projects = Project.objects.visible_to(user)
        tasks = Task.objects.visible_to(user)
        period_projects = projects.filter(created_at__date__gte=start, created_at__date__lte=end)
        period_tasks = tasks.filter(created_at__date__gte=start, created_at__date__lte=end)
        ranking_visible = SystemSetting.objects.filter(key="ranking_visible_to_staff", value="1").exists()
        can_view_ranking = not user.is_staff_role or ranking_visible
        refresh_staff_performance()
        ranking = StaffPerformance.objects.select_related("staff").order_by("-score", "-completion_rate")
        if user.is_staff_role:
            ranking = ranking.filter(staff=user)
        attendance = Attendance.objects.filter(date__gte=start, date__lte=end)
        if user.is_staff_role:
            attendance = attendance.filter(staff=user)
        elif user.is_manager_role:
            from apps.accounts.models import ManagerStaffRelation

            staff_ids = ManagerStaffRelation.objects.filter(manager=user).values_list("staff_id", flat=True)
            attendance = attendance.filter(staff_id__in=staff_ids)
        task_status_labels = dict(Task.Status.choices)
        attendance_status_labels = dict(Attendance.Status.choices)
        tasks_by_status = [
            {"label": task_status_labels.get(item["status"], item["status"]), "total": item["total"]}
            for item in tasks.values("status").annotate(total=Count("id"))
        ]
        attendance_summary = [
            {"label": attendance_status_labels.get(item["status"], item["status"]), "total": item["total"]}
            for item in attendance.values("status").annotate(total=Count("id"))
        ]
        ctx.update({
            "period": period,
            "period_label": label,
            "start": start,
            "end": end,
            "projects_total": projects.count(),
            "projects_done": projects.filter(status=Project.Status.DONE).count(),
            "period_projects": period_projects.count(),
            "tasks_total": tasks.count(),
            "tasks_done": tasks.filter(status=Task.Status.DONE).count(),
            "tasks_overdue": tasks.filter(deadline__lt=timezone.localdate()).exclude(status=Task.Status.DONE).count(),
            "period_tasks": period_tasks.count(),
            "tasks_by_status": tasks_by_status,
            "attendance_summary": attendance_summary,
            "ranking": ranking[:20] if can_view_ranking else [],
            "can_view_ranking": can_view_ranking,
            "ranking_visible_to_staff": ranking_visible,
        })
        return ctx


class RankingVisibilityView(LoginRequiredMixin, View):
    def post(self, request):
        if not request.user.is_admin_role:
            return HttpResponse(status=403)
        value = "1" if request.POST.get("ranking_visible_to_staff") == "on" else "0"
        SystemSetting.objects.update_or_create(
            key="ranking_visible_to_staff",
            defaults={"value": value, "description": "Cho nhân viên xem bảng xếp hạng"},
        )
        from django.shortcuts import redirect

        return redirect("reports")


def excel_response(wb, filename):
    output = BytesIO()
    wb.save(output)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


class ProjectExcelView(LoginRequiredMixin, View):
    def get(self, request):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Domain", "Manager", "Nhân viên", "Trạng thái", "Tiến độ", "Deadline", "Ngày tạo"])
        for p in Project.objects.visible_to(request.user).select_related("manager", "staff"):
            ws.append([p.domain, str(p.manager or ""), str(p.staff or ""), p.get_status_display(), p.progress, p.deadline, p.created_at])
        return excel_response(wb, "du-an.xlsx")


class TaskExcelView(LoginRequiredMixin, View):
    def get(self, request):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Công việc", "Người nhận", "Trạng thái", "Tiến độ", "Deadline"])
        for t in Task.objects.visible_to(request.user).select_related("assignee"):
            ws.append([t.title, str(t.assignee or ""), t.get_status_display(), t.progress, t.deadline])
        return excel_response(wb, "cong-viec.xlsx")


class PerformanceExcelView(LoginRequiredMixin, View):
    def get(self, request):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Nhân viên", "Được giao", "Hoàn thành", "Quá hạn", "Cần sửa", "Vướng mắc", "Tỷ lệ", "Điểm"])
        qs = StaffPerformance.objects.select_related("staff")
        if request.user.is_staff_role:
            qs = qs.filter(staff=request.user)
        for item in qs:
            ws.append([str(item.staff), item.total_assigned, item.total_done, item.overdue, item.returned_for_fix, item.blocked, item.completion_rate, item.score])
        return excel_response(wb, "hieu-suat.xlsx")


class LogsExcelView(LoginRequiredMixin, View):
    def get(self, request):
        if not request.user.is_admin_role:
            return HttpResponse(status=403)
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["User", "Hành động", "Loại", "ID", "IP", "Thời gian"])
        for log in ActivityLog.objects.select_related("actor")[:5000]:
            ws.append([str(log.actor or ""), log.display_action, log.object_type, log.object_id, log.ip_address, log.created_at])
        return excel_response(wb, "log-he-thong.xlsx")


class SummaryPdfView(LoginRequiredMixin, View):
    def get(self, request):
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="bao-cao.pdf"'
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas

        pdf = canvas.Canvas(response, pagesize=A4)
        pdf.drawString(60, 800, "Bao cao PhuLinhMedia")
        pdf.drawString(60, 770, f"Tong du an: {Project.objects.visible_to(request.user).count()}")
        pdf.drawString(60, 750, f"Tong cong viec: {Task.objects.visible_to(request.user).count()}")
        pdf.showPage()
        pdf.save()
        return response
